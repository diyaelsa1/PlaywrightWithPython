import openpyxl

class ExcelUtils:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path)

    def get_sheet(self, sheet_name):
        return self.workbook[sheet_name]

    def get_data_as_list_of_dicts(self, sheet_name):
        """
        Reads all rows and returns a list of dictionaries
        Each dictionary represents a row with keys from the header row
        """
        sheet = self.get_sheet(sheet_name)
        data = []
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        return data
